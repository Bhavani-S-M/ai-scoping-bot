# backend/app/routers/exports.py
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.config.database import get_async_session
from app.models.project_models import Project as ProjectModel
from app.auth.router import fastapi_users
from pydantic import BaseModel
import json
import os
import io
import zipfile
from datetime import datetime
from typing import Dict, Any

# For PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# For Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/exports", tags=["exports"])
current_active_user = fastapi_users.current_user(active=True)

class ExportRequest(BaseModel):
    project_id: str
    scope_data: Dict[str, Any]
    format: str  # 'pdf', 'excel', 'json', 'all'

# ============================================================================
# STEP 5: EXPORT FUNCTIONALITY
# ============================================================================

@router.post("/generate")
async def generate_export(
    export_request: ExportRequest,
    db: AsyncSession = Depends(get_async_session),
    user = Depends(current_active_user)
):
    """
    Generate exports in various formats
    Step 5: Download Professional Proposals
    """
    try:
        # Verify project access
        result = await db.execute(
            select(ProjectModel).where(
                ProjectModel.id == export_request.project_id,
                ProjectModel.owner_id == user.id
            )
        )
        project = result.scalar_one_or_none()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        
        export_format = export_request.format.lower()
        scope_data = export_request.scope_data
        
        # Create exports directory
        os.makedirs("exports", exist_ok=True)
        
        if export_format == 'pdf':
            file_path = await generate_pdf_export(project, scope_data)
            return FileResponse(
                file_path,
                media_type='application/pdf',
                filename=f"{project.name}_scope.pdf"
            )
        
        elif export_format == 'excel':
            file_path = await generate_excel_export(project, scope_data)
            return FileResponse(
                file_path,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=f"{project.name}_scope.xlsx"
            )
        
        elif export_format == 'json':
            file_path = await generate_json_export(project, scope_data)
            return FileResponse(
                file_path,
                media_type='application/json',
                filename=f"{project.name}_scope.json"
            )
        
        elif export_format == 'all':
            zip_path = await generate_all_exports(project, scope_data)
            return FileResponse(
                zip_path,
                media_type='application/zip',
                filename=f"{project.name}_scope_complete.zip"
            )
        
        else:
            raise HTTPException(status_code=400, detail="Invalid export format")
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export generation failed: {str(e)}")

# ============================================================================
# PDF EXPORT
# ============================================================================

async def generate_pdf_export(project: ProjectModel, scope_data: Dict[str, Any]) -> str:
    """Generate professional PDF document"""
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"exports/{project.name}_{timestamp}.pdf"
    
    doc = SimpleDocTemplate(filename, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title Page
    story.append(Paragraph(f"Project Scope Document", title_style))
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(f"<b>{project.name}</b>", styles['Heading1']))
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(f"Domain: {project.domain}", styles['Normal']))
    story.append(Paragraph(f"Complexity: {project.complexity}", styles['Normal']))
    story.append(PageBreak())
    
    # Overview Section
    overview = scope_data.get('overview', {})
    story.append(Paragraph("1. Project Overview", heading_style))
    story.append(Paragraph(overview.get('project_summary', 'N/A'), styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Key Objectives
    story.append(Paragraph("<b>Key Objectives:</b>", styles['Normal']))
    for obj in overview.get('key_objectives', []):
        story.append(Paragraph(f"• {obj}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Success Metrics
    story.append(Paragraph("<b>Success Metrics:</b>", styles['Normal']))
    for metric in overview.get('success_metrics', []):
        story.append(Paragraph(f"• {metric}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Timeline Section
    timeline = scope_data.get('timeline', {})
    story.append(Paragraph("2. Project Timeline", heading_style))
    story.append(Paragraph(f"Total Duration: {timeline.get('total_duration_months', 0)} months", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Phases Table
    phases = timeline.get('phases', [])
    if phases:
        phase_data = [['Phase', 'Duration (weeks)', 'Key Milestones']]
        for phase in phases:
            phase_data.append([
                phase.get('phase_name', ''),
                str(phase.get('duration_weeks', '')),
                ', '.join(phase.get('milestones', [])[:2])
            ])
        
        phase_table = Table(phase_data, colWidths=[2.5*inch, 1.5*inch, 3*inch])
        phase_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(phase_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Resources Section
    story.append(Paragraph("3. Resource Plan", heading_style))
    resources = scope_data.get('resources', [])
    
    if resources:
        resource_data = [['Role', 'Count', 'Effort (months)', 'Monthly Rate', 'Total Cost']]
        for resource in resources:
            resource_data.append([
                resource.get('role', ''),
                str(resource.get('count', '')),
                str(resource.get('effort_months', '')),
                f"${resource.get('monthly_rate', 0):,.0f}",
                f"${resource.get('total_cost', 0):,.0f}"
            ])
        
        resource_table = Table(resource_data, colWidths=[2*inch, 0.8*inch, 1.2*inch, 1.2*inch, 1.3*inch])
        resource_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(resource_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Cost Breakdown
    cost_breakdown = scope_data.get('cost_breakdown', {})
    story.append(Paragraph("4. Cost Summary", heading_style))
    story.append(Paragraph(f"<b>Total Project Cost:</b> ${cost_breakdown.get('total_cost', 0):,.2f}", styles['Normal']))
    story.append(Paragraph(f"<b>Contingency ({cost_breakdown.get('contingency_percentage', 15)}%):</b> ${cost_breakdown.get('contingency_amount', 0):,.2f}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Risks Section
    story.append(Paragraph("5. Key Risks & Mitigation", heading_style))
    risks = scope_data.get('risks', [])
    for risk in risks[:5]:  # Top 5 risks
        story.append(Paragraph(f"<b>{risk.get('risk', 'N/A')}</b> ({risk.get('severity', 'Medium')} severity)", styles['Normal']))
        story.append(Paragraph(f"Mitigation: {risk.get('mitigation', 'N/A')}", styles['Normal']))
        story.append(Spacer(1, 0.1*inch))
    
    # Build PDF
    doc.build(story)
    return filename

# ============================================================================
# EXCEL EXPORT
# ============================================================================

async def generate_excel_export(project: ProjectModel, scope_data: Dict[str, Any]) -> str:
    """Generate detailed Excel workbook"""
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"exports/{project.name}_{timestamp}.xlsx"
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Overview
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    overview = scope_data.get('overview', {})
    ws_overview['A1'] = "PROJECT OVERVIEW"
    ws_overview['A1'].font = Font(bold=True, size=16, color="1E40AF")
    
    row = 3
    ws_overview[f'A{row}'] = "Project Name:"
    ws_overview[f'B{row}'] = project.name
    ws_overview[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws_overview[f'A{row}'] = "Domain:"
    ws_overview[f'B{row}'] = project.domain
    
    row += 1
    ws_overview[f'A{row}'] = "Complexity:"
    ws_overview[f'B{row}'] = project.complexity
    
    row += 2
    ws_overview[f'A{row}'] = "Project Summary:"
    ws_overview[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_overview[f'A{row}'] = overview.get('project_summary', '')
    ws_overview[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    # Sheet 2: Timeline
    ws_timeline = wb.create_sheet("Timeline")
    ws_timeline['A1'] = "PROJECT TIMELINE"
    ws_timeline['A1'].font = Font(bold=True, size=14)
    
    timeline = scope_data.get('timeline', {})
    ws_timeline['A3'] = "Total Duration (months):"
    ws_timeline['B3'] = timeline.get('total_duration_months', 0)
    
    # Phases
    row = 5
    headers = ['Phase Name', 'Duration (weeks)', 'Start Week', 'End Week', 'Milestones']
    for col, header in enumerate(headers, 1):
        cell = ws_timeline.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    phases = timeline.get('phases', [])
    for phase in phases:
        row += 1
        ws_timeline.cell(row=row, column=1, value=phase.get('phase_name', ''))
        ws_timeline.cell(row=row, column=2, value=phase.get('duration_weeks', ''))
        ws_timeline.cell(row=row, column=3, value=phase.get('start_week', ''))
        ws_timeline.cell(row=row, column=4, value=phase.get('end_week', ''))
        ws_timeline.cell(row=row, column=5, value=', '.join(phase.get('milestones', [])))
        
        for col in range(1, 6):
            ws_timeline.cell(row=row, column=col).border = border
    
    # Sheet 3: Activities
    ws_activities = wb.create_sheet("Activities")
    ws_activities['A1'] = "PROJECT ACTIVITIES"
    ws_activities['A1'].font = Font(bold=True, size=14)
    
    row = 3
    headers = ['Activity Name', 'Phase', 'Effort (days)', 'Resources', 'Dependencies']
    for col, header in enumerate(headers, 1):
        cell = ws_activities.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    activities = scope_data.get('activities', [])
    for activity in activities:
        row += 1
        ws_activities.cell(row=row, column=1, value=activity.get('name', ''))
        ws_activities.cell(row=row, column=2, value=activity.get('phase', ''))
        ws_activities.cell(row=row, column=3, value=activity.get('effort_days', ''))
        ws_activities.cell(row=row, column=4, value=', '.join(activity.get('resources_needed', [])))
        ws_activities.cell(row=row, column=5, value=', '.join(activity.get('dependencies', [])))
        
        for col in range(1, 6):
            ws_activities.cell(row=row, column=col).border = border
    
    # Sheet 4: Resources & Costs
    ws_resources = wb.create_sheet("Resources & Costs")
    ws_resources['A1'] = "RESOURCE PLAN & COSTS"
    ws_resources['A1'].font = Font(bold=True, size=14)
    
    row = 3
    headers = ['Role', 'Count', 'Effort (months)', 'Allocation %', 'Monthly Rate', 'Total Cost']
    for col, header in enumerate(headers, 1):
        cell = ws_resources.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    resources = scope_data.get('resources', [])
    total_cost = 0
    for resource in resources:
        row += 1
        ws_resources.cell(row=row, column=1, value=resource.get('role', ''))
        ws_resources.cell(row=row, column=2, value=resource.get('count', ''))
        ws_resources.cell(row=row, column=3, value=resource.get('effort_months', ''))
        ws_resources.cell(row=row, column=4, value=resource.get('allocation_percentage', ''))
        ws_resources.cell(row=row, column=5, value=resource.get('monthly_rate', 0))
        ws_resources.cell(row=row, column=6, value=resource.get('total_cost', 0))
        
        # Format currency
        ws_resources.cell(row=row, column=5).number_format = '$#,##0.00'
        ws_resources.cell(row=row, column=6).number_format = '$#,##0.00'
        
        total_cost += resource.get('total_cost', 0)
        
        for col in range(1, 7):
            ws_resources.cell(row=row, column=col).border = border
    
    # Total row
    row += 1
    ws_resources.cell(row=row, column=5, value="TOTAL:")
    ws_resources.cell(row=row, column=5).font = Font(bold=True)
    ws_resources.cell(row=row, column=6, value=total_cost)
    ws_resources.cell(row=row, column=6).font = Font(bold=True)
    ws_resources.cell(row=row, column=6).number_format = '$#,##0.00'
    
    # Sheet 5: Risks
    ws_risks = wb.create_sheet("Risks")
    ws_risks['A1'] = "RISK REGISTER"
    ws_risks['A1'].font = Font(bold=True, size=14)
    
    row = 3
    headers = ['Risk', 'Category', 'Probability', 'Impact', 'Severity', 'Mitigation', 'Owner']
    for col, header in enumerate(headers, 1):
        cell = ws_risks.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    risks = scope_data.get('risks', [])
    for risk in risks:
        row += 1
        ws_risks.cell(row=row, column=1, value=risk.get('risk', ''))
        ws_risks.cell(row=row, column=2, value=risk.get('category', ''))
        ws_risks.cell(row=row, column=3, value=risk.get('probability', ''))
        ws_risks.cell(row=row, column=4, value=risk.get('impact', ''))
        ws_risks.cell(row=row, column=5, value=risk.get('severity', ''))
        ws_risks.cell(row=row, column=6, value=risk.get('mitigation', ''))
        ws_risks.cell(row=row, column=7, value=risk.get('owner', ''))
        
        for col in range(1, 8):
            ws_risks.cell(row=row, column=col).border = border
    
    # Auto-size columns
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    return filename

# ============================================================================
# JSON EXPORT
# ============================================================================

async def generate_json_export(project: ProjectModel, scope_data: Dict[str, Any]) -> str:
    """Generate structured JSON export"""
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"exports/{project.name}_{timestamp}.json"
    
    export_data = {
        "project_info": {
            "id": str(project.id),
            "name": project.name,
            "domain": project.domain,
            "complexity": project.complexity,
            "tech_stack": project.tech_stack,
            "created_at": project.created_at.isoformat()
        },
        "scope": scope_data,
        "exported_at": datetime.now().isoformat(),
        "version": "1.0"
    }
    
    with open(filename, 'w') as f:
        json.dump(export_data, f, indent=2)
    
    return filename

# ============================================================================
# ALL FORMATS (ZIP)
# ============================================================================

async def generate_all_exports(project: ProjectModel, scope_data: Dict[str, Any]) -> str:
    """Generate all formats and package in ZIP"""
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_filename = f"exports/{project.name}_{timestamp}_complete.zip"
    
    # Generate all formats
    pdf_file = await generate_pdf_export(project, scope_data)
    excel_file = await generate_excel_export(project, scope_data)
    json_file = await generate_json_export(project, scope_data)
    
    # Create ZIP
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(pdf_file, os.path.basename(pdf_file))
        zipf.write(excel_file, os.path.basename(excel_file))
        zipf.write(json_file, os.path.basename(json_file))
    
    # Clean up individual files
    os.remove(pdf_file)
    os.remove(excel_file)
    os.remove(json_file)
    
    return zip_filename